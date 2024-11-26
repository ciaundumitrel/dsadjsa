import calendar
import json

from ortools.sat.python import cp_model
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import customtkinter as ctk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
import pandas as pd


class NurseSchedulerApp:
    def __init__(self, config_path):
        self.config_path = config_path
        self.config = self.load_config()
        self.root = None
        self.month = None
        self.year = None
        self.nurses = []
        self.free_days = {}
        self.weekends = []
        self.days = None
        self.edited_data = []
        self.preloaded_data = {}
        self.preloaded_days = {}

    def load_config(self):
        try:
            with open(self.config_path, 'r') as f:
                return json.load(f)
        except FileNotFoundError:
            return {"month": 1, "year": 2024, "nurses": [], "free_days": {}}

    def load_excel(self):
        file_path = filedialog.askopenfilename(
            title="Select an Excel file",
            filetypes=[("Excel Files", "*.xlsx;*.xls")]
        )

        if file_path:
            try:
                df = pd.read_excel(file_path)
                nurses_schedule = {
                    row['Nurse']: row.drop('Nurse').to_dict()
                    for _, row in df.iterrows()
                }
                self.preloaded_data = {x: nurses_schedule[x] for x in nurses_schedule if x not in {'Days'}}
                self.preloaded_days = nurses_schedule['Days']
            except Exception as e:
                print(f"Error loading Excel file: {e}")

    def save_config(self):
        try:
            with open(self.config_path, 'w') as f:
                json.dump(self.config, f, indent=2)
            messagebox.showinfo("Success", "Configuration saved successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save configuration:\n{e}")

    def generate_schedule(self):
        self.nurses = self.config.get("nurses", None)
        free_days = self.config.get("free_days", None)
        self.month = self.config.get("month", None)
        self.year = self.config.get("year", None)

        if self.nurses is None or free_days is None or self.month is None or self.year is None:
            messagebox.showerror("Error", "Invalid config file.")

        shifts = ['L', 'Z', 'N', '8', 'CO']
        hours = {'L': 0, 'Z': 12, 'N': 12, '8': 8, 'CO': 8}

        starting_day = calendar.monthrange(self.year, self.month)[0]
        self.days = calendar.monthrange(self.year, self.month)[1]

        for i in range(self.days):
            if (starting_day + i) % 7 == 5:
                self.weekends.append(i)
            elif (starting_day + i) % 7 == 6:
                self.weekends.append(i)

        required_hours = self.days * 8 - len(self.weekends) * 8
        model = cp_model.CpModel()

        x = {}

        for n in range(len(self.nurses)):
            for d in range(self.days):
                for s in shifts:
                    x[n, d, s] = model.NewBoolVar(f'x_{n}_{d}_{s}')

        for index, nurse in enumerate(self.preloaded_data.keys()):
            for shift in self.preloaded_data[nurse]:
                shift_value = self.preloaded_data[nurse][shift]
                if shift_value in ['N', 'Z', '8']:
                    model.Add(x[index, int(shift)-1, shift_value] == 1)
                    model.Add(x[index, int(shift) - 1, 'CO'] == 0)
                elif shift_value == 'CO':
                    if shift in self.weekends:
                        model.Add(x[index, int(shift)-1, 'L'] == 1)
                        model.Add(x[index, int(shift)-1, 'CO'] == 0)
                    else:
                        model.Add(x[index, int(shift)-1, 'CO'] == 1)
                else:
                    model.Add(x[index, int(shift) - 1, 'CO'] == 0)

        for n in range(len(self.nurses)):
            model.Add(
                sum(x[n, d, s] * hours[s] for d in range(self.days) for s in shifts) == required_hours
            )
            for d in range(self.days):
                model.Add(sum(x[n, d, s] for s in shifts) == 1)
            model.Add(
                sum(x[n, d, 'N'] for d in range(self.days)) >= 3
            )
            model.Add(
                sum(x[n, d, 'Z'] for d in range(self.days)) >= 3
            )
            model.Add(
                sum(x[n, d, '8'] for d in range(self.days)) < 5
            )
            for d in range(self.days - 1):
                model.AddImplication(x[n, d, 'N'], x[n, d + 1, 'L'])

        for n in range(len(self.nurses)):
            for d in range(self.days - 1):
                next_is_n_or_l = model.NewBoolVar(f'next_is_n_or_l_{n}_{d}')
                model.Add(x[n, d + 1, 'N'] + x[n, d + 1, 'L'] == 1).OnlyEnforceIf(next_is_n_or_l)
                model.Add(x[n, d + 1, 'N'] + x[n, d + 1, 'L'] != 1).OnlyEnforceIf(next_is_n_or_l.Not())
                model.AddImplication(x[n, d, 'Z'], next_is_n_or_l)

        for d in range(self.days):
            if (starting_day + d) % 7 in [0, 1, 2, 3, 4]:
                model.Add(
                    sum(x[n, d, 'Z'] + x[n, d, '8'] for n in range(len(self.nurses))) >= 4
                )
                model.Add(
                    sum(x[n, d, 'Z'] + x[n, d, '8'] for n in range(len(self.nurses))) <= 7
                )
                model.Add(
                    sum(x[n, d, '8'] for n in range(len(self.nurses))) > 0
                )
                model.Add(
                    sum(x[n, d, '8'] for n in range(len(self.nurses))) < 5
                )

                model.Add(
                    sum(x[n, d, 'Z'] for n in range(len(self.nurses))) >= 2
                )
                model.Add(
                    sum(x[n, d, 'Z'] for n in range(len(self.nurses))) <= 4
                )
                model.Add(
                    sum(x[n, d, 'N'] for n in range(len(self.nurses))) == 2
                )

        for d in self.weekends:
            model.Add(
                sum(x[n, d, 'Z'] for n in range(len(self.nurses))) == 2
            )
            model.Add(
                sum(x[n, d, 'N'] for n in range(len(self.nurses))) == 2
            )
            model.Add(
                sum(x[n, d, '8'] for n in range(len(self.nurses))) == 0
            )

        solver = cp_model.CpSolver()
        status = solver.Solve(model)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Nurse Schedule"

        header = ["Nurse"] + [f"{d + 1}" for d in range(self.days)]
        sheet.append(header)

        day_initials = ['L', 'M', 'M', 'J', 'V', 'S', 'D'] * 6

        sheet.insert_rows(2)
        for d, day in enumerate(day_initials[starting_day:self.days + starting_day], start=2):
            sheet.cell(row=2, column=d).value = day

        if status == cp_model.OPTIMAL:
            for n in range(len(self.nurses)):
                row = [self.nurses[n]]
                total_hours = 0
                for d in range(self.days):
                    shift_assigned = None

                    for s in shifts:
                        if solver.Value(x[n, d, s]) == 1:
                            shift_assigned = s
                            total_hours += hours[s]
                            break

                    if shift_assigned == 'L':
                        row.append("")
                    else:
                        row.append(shift_assigned)

                row.append(total_hours)
                sheet.append(row)

            shift_counts = {shift: [0] * self.days for shift in shifts}
            for n in range(len(self.nurses)):
                for d in range(self.days):
                    for s in shifts:
                        if solver.Value(x[n, d, s]) == 1:
                            shift_counts[s][d] += 1

            summary_start_row = len(self.nurses) + 4
            sheet.append(["Shift Summary"] + [""] * self.days)
            sheet.merge_cells(start_row=summary_start_row - 1, start_column=1, end_row=summary_start_row - 1,
                              end_column=self.days + 1)
            for shift, counts in shift_counts.items():
                sheet.append([shift] + counts)

            weekend_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

            for d in self.weekends:
                col_idx = d + 2
                for row in sheet.iter_rows(min_row=2, max_row=len(self.nurses) + 3):
                    cell = row[col_idx - 1]
                    cell.fill = weekend_fill

            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                sheet.column_dimensions[column].width = max_length + 2

            messagebox.showinfo(
                "Schedule Generated",
                f"Schedule for {self.config['month']}/{self.config['year']} generated.",
            )
            print(f"Configuration used:\n{json.dumps(self.config, indent=2)}")
            workbook.save(f"nurse_schedule_{self.month}_{self.year}.xlsx")
            self.weekends = []

            return 1
        else:
            messagebox.showerror(
                "Schedule Generation Failed",
                "No solution found",
            )
            return 0

    def generate_empty_excel(self):
        self.nurses = self.config.get("nurses", None)
        self.month = self.config.get("month", None)
        self.year = self.config.get("year", None)

        self.days = calendar.monthrange(self.year, self.month)[1]

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Nurse Schedule"
        starting_day = calendar.monthrange(self.year, self.month)[0]
        header = ["Nurse"] + [f"{d + 1}" for d in range(self.days)]
        sheet.append(header)
        grey_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

        day_initials = ['L', 'M', 'M', 'J', 'V', 'S', 'D'] * 6

        sheet.insert_rows(2)
        sheet.cell(row=2, column=1).value = 'Days'
        for d, day in enumerate(day_initials[starting_day:self.days + starting_day], start=2):
            sheet.cell(row=2, column=d).value = day


        for n in range(len(self.nurses)):
            row = [self.nurses[n]]
            for d in range(self.days):
                row.append("")
            sheet.append(row)


        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[column].width = max_length + 2



        file_name = f"EMPTY_nurse_schedule_{self.month}_{self.year}.xlsx"
        workbook.save(file_name)
        messagebox.showinfo(
            "Schedule Generated",
            f"EMPTY Schedule for {self.month}/{self.year} generated.",
        )
        edited_workbook = load_workbook(file_name)
        edited_sheet = edited_workbook.active

        for row in edited_sheet.iter_rows(values_only=True):
            self.edited_data.append(row)

    def create_main_view(self, tab):
        frame = ctk.CTkFrame(tab)
        frame.pack(padx=10, pady=10, fill="both", expand=True)

        ctk.CTkLabel(frame, text="Nurse Scheduler").pack(pady=10)

        load_excel = ctk.CTkButton(
            frame, text="Load Excel", command=self.load_excel
        )
        load_excel.pack(pady=20)

        generate_button = ctk.CTkButton(
            frame, text="Generate Schedule ", command=self.generate_schedule
        )
        generate_button.pack(pady=20)

    def create_generate_empty_view(self, tab):
        frame = ctk.CTkFrame(tab)
        frame.pack(padx=10, pady=10, fill="both", expand=True)

        ctk.CTkLabel(frame, text="Nurse Scheduler").pack(pady=10)

        generate_empty_excel_button = ctk.CTkButton(
            frame, text="Generate Empty Excel", command=self.generate_empty_excel
        )
        generate_empty_excel_button.pack(pady=20)

    def create_config_view(self, tab):
        frame = ctk.CTkFrame(tab)
        frame.pack(padx=10, pady=10, fill="both", expand=True)

        ctk.CTkLabel(frame, text="Month:").grid(row=0, column=0, padx=5, pady=5)
        month_var = ctk.IntVar(value=self.config.get("month", 1))
        ctk.CTkEntry(frame, textvariable=month_var, width=50).grid(row=0, column=1, padx=5, pady=5)

        ctk.CTkLabel(frame, text="Year:").grid(row=0, column=2, padx=5, pady=5)
        year_var = ctk.IntVar(value=self.config.get("year", 2024))
        ctk.CTkEntry(frame, textvariable=year_var, width=80).grid(row=0, column=3, padx=5, pady=5)

        ctk.CTkLabel(frame, text="Nurses:").grid(row=1, column=0, padx=5, pady=5, sticky="nw")
        nurses_listbox = ctk.CTkTextbox(frame, width=200, height=250)
        nurses_listbox.grid(row=1, column=1, columnspan=3, padx=5, pady=5)
        nurses_listbox.insert("1.0", "\n".join(self.config.get("nurses", [])))

        def save_changes():
            self.config["month"] = month_var.get()
            self.config["year"] = year_var.get()
            self.config["nurses"] = nurses_listbox.get("1.0", "end").strip().split("\n")
            self.save_config()

        save_button = ctk.CTkButton(frame, text="Save Changes", command=save_changes)
        save_button.grid(row=3, column=1, columnspan=2, pady=10)

    def create_gui(self):
        ctk.set_appearance_mode("System")
        ctk.set_default_color_theme("blue")

        self.root = ctk.CTk()
        self.root.title("Nurse Scheduler")
        self.root.geometry("600x500")

        tabview = ctk.CTkTabview(self.root)
        tabview.pack(fill="both", expand=True, padx=10, pady=10)

        generate_excel_from_config = tabview.add("Main")
        generate_excel_from_empty = tabview.add("Generate From Empty")
        config_tab = tabview.add("Configuration")

        self.create_main_view(generate_excel_from_config)
        self.create_config_view(config_tab)
        self.create_generate_empty_view(generate_excel_from_empty)
        self.root.mainloop()


if __name__ == "__main__":
    app = NurseSchedulerApp('config.json')
    app.create_gui()
